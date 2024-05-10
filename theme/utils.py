import itertools
import os
import urllib
from functools import wraps
from io import BytesIO
from os.path import splitext
from typing import Dict, IO, Union, List, Iterator
from urllib.parse import urlparse, ParseResult, parse_qs, urlencode, urlunparse
from urllib.request import urlopen
from uuid import uuid4

import pandas as pd
from PIL import Image
from django.conf import settings
from django.contrib.auth import REDIRECT_FIELD_NAME
from django.contrib.auth.decorators import login_required as django_login_required
from django.core.files import File
from django.core.files.base import ContentFile
from django.db import models
from django.http import HttpResponseRedirect, FileResponse
from django.shortcuts import resolve_url
from django.utils import timezone
from django.utils.encoding import force_str
from django_htmx.http import HttpResponseClientRedirect
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def uuid_name_upload_to(instance: models.Model, filename: str) -> str:
    app_label = instance.__class__._meta.app_label
    cls_name = instance.__class__.__name__.lower()
    ymd_path = force_str(timezone.now().strftime("%Y/%m/%d"))
    extension = splitext(filename)[-1].lower()
    new_filename = uuid4().hex + extension
    return "/".join((app_label, cls_name, ymd_path, new_filename))


def make_thumb(
    image_file: File, max_width: int = 1024, max_height: int = 1024, quality=100
) -> File:
    pil_image = Image.open(image_file)
    max_size = (max_width, max_height)
    pil_image.thumbnail(max_size)
    if pil_image.mode != "RGB":
        pil_image = pil_image.convert("RGB")

    thumb_name = splitext(image_file.name)[0] + ".jpg"
    thumb_file = ContentFile(b"", name=thumb_name)
    pil_image.save(thumb_file, format="jpeg", quality=quality)

    return thumb_file


def login_required_hx(
    function=None,
    redirect_field_name=REDIRECT_FIELD_NAME,
    login_url=None,
):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            decorated_view_func = django_login_required(
                function=view_func,
                redirect_field_name=redirect_field_name,
                login_url=login_url,
            )

            response = decorated_view_func(request, *args, **kwargs)

            if isinstance(response, HttpResponseRedirect):
                resolved_login_url = resolve_url(login_url or settings.LOGIN_URL)

                if request.htmx and resolved_login_url in response.url:
                    # /accounts/login/?next=/blog/tags/new/%3F_%3D1710826915601
                    # next_url: str = response.url
                    # request.htmx.current_url
                    # return HttpResponseClientRedirect(next_url)

                    # HTMX 요청을 한 주소가 next 인자로 포함된 URL
                    redirect_url: str = response.url
                    # HTMX 요청을 한 페이지의 주소
                    new_redirect_url: str = request.htmx.current_url

                    # next 인자가 포함된 URL에서 next 인자만 new_redirect_url 값으로 변경
                    parsed: ParseResult = urlparse(redirect_url)
                    query_dict: dict = parse_qs(parsed.query)
                    query_dict["next"] = [new_redirect_url]
                    new_query: str = urlencode(query_dict, doseq=True)
                    new_next_url = urlunparse(
                        (
                            parsed.scheme,
                            parsed.netloc,
                            parsed.path,
                            parsed.params,
                            new_query,
                            parsed.fragment,
                        )
                    )

                    return HttpResponseClientRedirect(new_next_url)

            return response

        return wrapper

    if function:
        return decorator(function)

    return decorator


def generate_excel_response(data: Union[Dict, List], filename: str) -> FileResponse:
    excel_file = create_excel_file(data, filename)
    response = FileResponse(
        excel_file,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def calculate_dimension(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2) * 1.2  # 조정된 폭 계산
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def cell_pattern_fill(
    df: pd.DataFrame,
    worksheet: Worksheet,
    head_fill_color: str,
    head_font_color: str,
    body_fill_color: str,
    body_font_color: str,
    fill_type: str,
) -> None:
    # 최상위 행
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = PatternFill(
            start_color=head_fill_color,
            end_color=head_fill_color,
            fill_type=fill_type,  # noqa
        )
        cell.font = Font(color=head_font_color, bold=True)

    # 하위 행
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = PatternFill(
                    start_color=body_fill_color,
                    end_color=body_fill_color,
                    fill_type=fill_type,  # noqa
                )
                cell.font = Font(color=body_font_color)
            if col == 4:
                cell.hyperlink = str(df.iloc[row - 2]["thumb"])
                cell.font = Font(color="0000FF")


def create_excel_file(data: Union[Dict, List], filename: str) -> IO[bytes]:
    df = pd.json_normalize(data)

    if "url" in df.columns:
        df = df.drop("url", axis=1)

    io = BytesIO()
    io.name = filename
    writer = pd.ExcelWriter(io, engine="openpyxl")  # noqa
    df.to_excel(writer, index=False, sheet_name="Sheet1")
    workbook = writer.book
    worksheet = workbook.active

    cell_pattern_fill(
        df,
        worksheet,
        "4472c4",
        "FFFFFF",
        "d9e1f2",
        "000000",
        "solid",
    )
    calculate_dimension(worksheet)

    writer._save()  # noqa
    io.seek(0)
    return io


def get_chunks(iterable: Iterator, chunk_size: int = 100) -> Iterator:
    iterator = iterable if hasattr(iterable, "__next__") else iter(iterable)
    for first in iterator:
        yield itertools.chain([first], itertools.islice(iterator, chunk_size - 1))


def convert_file(url: str):
    response = urllib.request.urlopen(url)
    filename = os.path.basename(urlparse(url).path)
    image_file = ContentFile(content=response.read(), name=filename)
    return image_file
